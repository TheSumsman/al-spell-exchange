"""Build build/SpellExchange.xlsx from data/wizard-spells.csv.

The workbook is authored as .xlsx but LIVES IN GOOGLE SHEETS, so every formula
is restricted to the intersection of both dialects: COUNTIF(S), SUMIF, TEXTJOIN,
INDEX, IF, SEARCH, CEILING -- filled down explicitly.

Deliberately avoided:
  * QUERY / ARRAYFORMULA / FILTER      -- Google-only
  * spilling dynamic arrays / XLOOKUP  -- Excel-365-only, converts badly
  * checkbox data validation           -- does not survive xlsx -> Sheets
"""
import csv
import os
import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

sys.stdout.reconfigure(encoding="utf-8")

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
CSV_PATH = os.path.join(ROOT, "data", "wizard-spells.csv")
OUT = os.path.join(ROOT, "build", "SpellExchange.xlsx")

N_WIZ = 24          # pre-sized roster
FIRST_RESP = 2      # first data row on the response tab
PLAN_TOP = 11       # first spell row on Copy Planner

# The placeholder tab this workbook ships. Google Forms always creates its OWN
# tab when a form is linked (typically "Form Responses 1"), so after linking you
# repoint the formulas at Google's tab with ONE find-and-replace and then delete
# this one. Do NOT name it "Form Responses 1" -- Google would collide with it and
# fall back to "Form Responses 2", which is just confusing.
RESP_TAB = "Form Responses"

# Bound the response ranges rather than using whole columns ($D:$D). Whole
# columns are equally insertion-safe but make the offline formula evaluator
# unusably slow, and 200 rows is far more headroom than a 24-wizard roster
# needs. Past ~199 responses the workbook would need rebuilding anyway.
RESP_LAST = 200


def resp(col, row):
    """A reference to the response tab that survives new form submissions.

    Google Forms INSERTS a row for each response instead of filling a blank
    one, and the insert lands exactly where a direct reference points -- so
    'Form Responses'!D2 silently becomes D3, D4, D5... one row per submission,
    and the whole workbook drifts off the data.

    INDEX over the WHOLE column is immune: inserting rows never rewrites a
    full-column range, and the row index is a plain number, not a reference.
    """
    return "INDEX('%s'!$%s$1:$%s$%d,%d)" % (RESP_TAB, col, col, RESP_LAST, row)

HDR_FILL = PatternFill("solid", fgColor="2F3E46")
HDR_FONT = Font(color="FFFFFF", bold=True)
NOTE_FONT = Font(italic=True, color="555555")
TITLE_FONT = Font(bold=True, size=14)
BOX_FILL = PatternFill("solid", fgColor="EDF2F4")
THIN = Side(style="thin", color="BBBBBB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def header(ws, row, labels, widths=None):
    for i, lab in enumerate(labels, start=1):
        c = ws.cell(row=row, column=i, value=lab)
        c.fill, c.font = HDR_FILL, HDR_FONT
        c.alignment = Alignment(vertical="center", wrap_text=True)
    if widths:
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w


def load_spells():
    with open(CSV_PATH, encoding="utf-8") as fh:
        rows = list(csv.DictReader(fh))
    rows.sort(key=lambda r: (int(r["Level"]), r["Name"].lower()))
    return rows


def build(spells=None, n_wiz=None, out=None):
    """Build the workbook. Parameterised so the verifier can build a small one."""
    global N_WIZ
    spells = load_spells() if spells is None else spells
    if n_wiz is not None:
        N_WIZ = n_wiz
    out = out or OUT
    n = len(spells)
    wb = Workbook()

    # ------------------------------------------------------------ Read Me
    ws = wb.active
    ws.title = "Read Me"
    ws.column_dimensions["A"].width = 108
    lines = [
        ("Wizard Spell Exchange - AL Epic (Forgotten Realms)", TITLE_FONT),
        ("", None),
        ("Players register their spellbook via the Google Form. Everything else on this "
         "workbook is calculated - do not type into any tab except the yellow cells on "
         "Copy Planner.", None),
        ("", None),
        ("HOW TO USE", TITLE_FONT),
        ("1. Fill in the Google Form at the event (one submission per wizard character).", None),
        ("2. Open the 'Copy Planner' tab and pick your character in the yellow cell.", None),
        ("3. Filter the Status column to 'CAN COPY' to see what is available to you.", None),
        ("4. Tick the 'Want' checkbox. The totals at the top add up GP and Downtime.", None),
        ("5. Open 'Copy Log' for the list of what you copied, what it cost, and from", None),
        ("   whom - plus a line of text ready to paste into your character log.", None),
        ("", None),
        ("THE COST", TITLE_FONT),
        ("50 GP per spell level, and 1 Downtime Day per spell for spell levels 1-4, "
         "2 DT per spell for levels 5-9.", None),
        ("", None),
        ("PHB 2024, Wizard - 'Expanding and Replacing a Spellbook':", Font(bold=True)),
        ('  "When you find a level 1+ Wizard spell, you can copy it into your spellbook if '
         "it's of a level you can", NOTE_FONT),
        ('   prepare and if you have time to copy it. For each level of the spell, the '
         'transcription takes 2 hours', NOTE_FONT),
        ('   and costs 50 GP."', NOTE_FONT),
        ("", None),
        ("ALPG v2026.4 p.3, Downtime - 'Copying Spells':", Font(bold=True)),
        ('  "Use \'Expanding and Replacing a Spellbook\' (PH) to copy spells found in '
         "adventures at 1 DT per spell", NOTE_FONT),
        ("   up to level 4 and 2 DT per spell at levels 5-9. You may copy spells from a "
         "character's spellbook", NOTE_FONT),
        ('   immediately after a session in which you both played."', NOTE_FONT),
        ("", None),
        ("ALPG p.2 - Order of Scribes wizards copy ten level 1-4 spells, or five level 5-9 "
         "spells, for 1 DT.", None),
        ("This workbook applies that rate automatically when the subclass is recorded.", NOTE_FONT),
        ("", None),
        ("GOLD AND DOWNTIME", TITLE_FONT),
        ("Both matter, so this workbook totals both. Downtime is the one people forget: "
         "you earn 10 DT", None),
        ("per session (ALPG p.6), levelling up costs 10 DT and a Bastion turn 7 DT.", None),
        ("", None),
        ("The Downtime budget on Copy Planner starts at 10 - the minimum you are sure to "
         "have after", None),
        ("this Epic. If your character has downtime banked in their log, change that cell "
         "to your real total.", NOTE_FONT),
        ("", None),
        ("RULING FOR THIS EVENT", TITLE_FONT),
        ("* The whole Epic counts as one session: any wizard present may copy from any "
         "other wizard present.", None),
        ("  ALPG does not address multi-table Epics, so this fills a genuine gap.", NOTE_FONT),
        ("", None),
        ("Everything else here is the rules as written, quoted above - 50 GP per spell "
         "level, and you may", None),
        ("only copy a spell of a level you can already prepare. Neither is a judgement "
         "call.", None),
        ("", None),
        ("If a player cites the PHB's 'Copying the Book' clause at 10 GP per level: that "
         "covers duplicating", NOTE_FONT),
        ("your own spellbook into a replacement, not learning a spell from another "
         "wizard. It does not apply.", NOTE_FONT),
        ("", None),
        ("Spell list: %d leveled wizard spells from the AL-legal Forgotten Realms sources. "
         "Cantrips are excluded" % n, NOTE_FONT),
        ("- they are never kept in a spellbook.", NOTE_FONT),
    ]
    for i, (text, font) in enumerate(lines, start=1):
        c = ws.cell(row=i, column=1, value=text)
        if font:
            c.font = font
        c.alignment = Alignment(wrap_text=False)

    # ------------------------------------------------------------- Spells
    sp = wb.create_sheet("Spells")
    header(sp, 1, ["Level", "Spell", "School", "Source", "Restriction", "GP", "DT"],
           [7, 34, 15, 22, 22, 8, 6])
    for i, r in enumerate(spells, start=2):
        sp.cell(row=i, column=1, value=int(r["Level"]))
        sp.cell(row=i, column=2, value=r["Name"])
        sp.cell(row=i, column=3, value=r["School"])
        sp.cell(row=i, column=4, value=r["Source"])
        sp.cell(row=i, column=5, value=r["Restriction"])
        sp.cell(row=i, column=6, value="=A%d*50" % i)
        sp.cell(row=i, column=7, value="=IF(A%d<=4,1,2)" % i)
    sp.freeze_panes = "A2"
    sp.auto_filter.ref = "A1:G%d" % (n + 1)

    # ----------------------------------------------------- Form Responses
    fr = wb.create_sheet(RESP_TAB)
    fr_cols = ["Timestamp", "Email Address", "Player name", "Character name",
               "Wizard level", "Table number", "Wizard subclass",
               "Contact after the event"]
    fr_cols += ["Level %d spells" % i for i in range(1, 10)]
    fr_cols += ["Other spells not in the lists above"]
    header(fr, 1, fr_cols, [18, 24, 18, 20, 12, 12, 22, 20] + [30] * 10)
    fr.cell(row=N_WIZ + 3, column=1,
            value="Google Forms writes into this tab. Do not edit or reorder columns - "
                  "the whole workbook references them by position.").font = NOTE_FONT
    fr.freeze_panes = "C2"

    # ------------------------------------------------------------ Wizards
    wz = wb.create_sheet("Wizards")
    header(wz, 1, ["Character", "Player", "Wizard level", "Max spell level", "Table",
                   "Subclass", "Order of Scribes?", "Contact", "Spells in book",
                   "Rare spells held"],
           [22, 18, 12, 13, 8, 22, 15, 20, 13, 13])
    for k in range(1, N_WIZ + 1):
        r = k + 1
        src = FIRST_RESP + k - 1
        wz.cell(row=r, column=1,
                value="=IF(%s=\"\",\"\",%s)" % (resp("D", src), resp("D", src)))
        wz.cell(row=r, column=2, value="=IF($A%d=\"\",\"\",%s)" % (r, resp("C", src)))
        wz.cell(row=r, column=3, value="=IF($A%d=\"\",\"\",%s)" % (r, resp("E", src)))
        # Wizard slot levels arrive at levels 1,3,5,...,17 -> MIN(9, roundup(lvl/2))
        wz.cell(row=r, column=4, value="=IF($A%d=\"\",\"\",MIN(9,CEILING($C%d/2,1)))" % (r, r))
        wz.cell(row=r, column=5, value="=IF($A%d=\"\",\"\",%s)" % (r, resp("F", src)))
        wz.cell(row=r, column=6, value="=IF($A%d=\"\",\"\",%s)" % (r, resp("G", src)))
        wz.cell(row=r, column=7,
                value="=IF($A%d=\"\",\"\",IF(ISNUMBER(SEARCH(\"Scribes\",%s)),\"Yes\",\"No\"))"
                      % (r, resp("G", src)))
        wz.cell(row=r, column=8, value="=IF($A%d=\"\",\"\",%s)" % (r, resp("H", src)))
        col = get_column_letter(3 + k)          # this wizard's column on Matrix
        wz.cell(row=r, column=9,
                value="=IF($A%d=\"\",\"\",COUNT(Matrix!$%s$2:$%s$%d))" % (r, col, col, n + 1))
        # spells this wizard holds that nobody else has
        wz.cell(row=r, column=10,
                value="=IF($A%d=\"\",\"\",COUNTIFS(Matrix!$%s$2:$%s$%d,1,Matrix!$%s$2:$%s$%d,1))"
                      % (r, col, col, n + 1,
                         get_column_letter(3 + N_WIZ + N_WIZ + 2),
                         get_column_letter(3 + N_WIZ + N_WIZ + 2), n + 1))
    wz.freeze_panes = "A2"

    # ------------------------------------------------------------- Matrix
    mx = wb.create_sheet("Matrix")
    name_first = 4                                  # column D
    name_last = 3 + N_WIZ                           # column AA
    help_first = name_last + 1                      # helper name block
    help_last = help_first + N_WIZ - 1
    own_col = help_last + 1                         # Owners
    cnt_col = own_col + 1                           # # Owners

    header(mx, 1, ["Level", "Spell", "School"], [7, 34, 15])
    for k in range(1, N_WIZ + 1):
        c = mx.cell(row=1, column=name_first + k - 1,
                    value="=IF(Wizards!A%d=\"\",\"\",Wizards!A%d)" % (k + 1, k + 1))
        c.fill, c.font = HDR_FILL, HDR_FONT
        mx.column_dimensions[get_column_letter(name_first + k - 1)].width = 14
        h = mx.cell(row=1, column=help_first + k - 1, value="helper %d" % k)
        h.fill, h.font = HDR_FILL, HDR_FONT
        mx.column_dimensions[get_column_letter(help_first + k - 1)].width = 14
        mx.column_dimensions[get_column_letter(help_first + k - 1)].hidden = True
    for col, lab, w in ((own_col, "Owners", 46), (cnt_col, "# Owners", 10)):
        c = mx.cell(row=1, column=col, value=lab)
        c.fill, c.font = HDR_FILL, HDR_FONT
        mx.column_dimensions[get_column_letter(col)].width = w

    for i, r in enumerate(spells, start=2):
        mx.cell(row=i, column=1, value=int(r["Level"]))
        mx.cell(row=i, column=2, value=r["Name"])
        mx.cell(row=i, column=3, value=r["School"])
        for k in range(1, N_WIZ + 1):
            src = FIRST_RESP + k - 1
            # The response cell for this spell's level: columns I..Q = levels 1..9.
            # Whole-column INDEX for the same insertion-safety reason as resp().
            lvl_cell = ("INDEX('%s'!$I$1:$Q$%d,%d,$A%d)"
                        % (RESP_TAB, RESP_LAST, src, i))
            other = resp("R", src)
            # Delimiter-wrapped so "Fire Bolt" never matches inside "Wall of Fire".
            test = ('OR(ISNUMBER(SEARCH(", "&$B{r}&", ", ", "&{lv}&", ")),'
                    'ISNUMBER(SEARCH(", "&$B{r}&", ", ", "&{ot}&", ")))'
                    ).format(r=i, lv=lvl_cell, ot=other)
            mx.cell(row=i, column=name_first + k - 1,
                    value="=IF(%s=\"\",\"\",IF(%s,1,\"\"))"
                          % ("Wizards!$A$%d" % (k + 1), test))
            nm = get_column_letter(name_first + k - 1)
            mx.cell(row=i, column=help_first + k - 1,
                    value="=IF(%s%d=1,%s$1,\"\")" % (nm, i, nm))
        mx.cell(row=i, column=own_col,
                value="=TEXTJOIN(\", \",TRUE,$%s%d:$%s%d)"
                      % (get_column_letter(help_first), i,
                         get_column_letter(help_last), i))
        mx.cell(row=i, column=cnt_col,
                value="=COUNT($%s%d:$%s%d)"
                      % (get_column_letter(name_first), i,
                         get_column_letter(name_last), i))
    mx.freeze_panes = "D2"

    # -------------------------------------------------------- Copy Planner
    cp = wb.create_sheet("Copy Planner")
    cp.column_dimensions["A"].width = 7
    cp.column_dimensions["B"].width = 34
    cp.column_dimensions["C"].width = 15
    cp.column_dimensions["D"].width = 22
    cp.column_dimensions["E"].width = 16
    cp.column_dimensions["F"].width = 44
    cp.column_dimensions["G"].width = 8
    cp.column_dimensions["H"].width = 6
    cp.column_dimensions["I"].width = 8

    cp["A1"] = "Your character:"
    cp["A1"].font = Font(bold=True)
    pick = cp["B1"]
    pick.fill = PatternFill("solid", fgColor="FFF3B0")
    pick.border = BORDER
    cp["C1"] = "<- pick from the list, then filter Status to CAN COPY"
    cp["C1"].font = NOTE_FONT

    cp["A2"] = "Downtime budget (DT):"
    cp["A2"].font = Font(bold=True)
    bud = cp["B2"]
    bud.value = 10
    bud.fill = PatternFill("solid", fgColor="FFF3B0")
    bud.border = BORDER
    cp["C2"] = ("10 DT is the minimum you are sure to have after this Epic - "
                "change it to your character's real total if more is banked")
    cp["C2"].font = NOTE_FONT

    last = PLAN_TOP + n - 1
    # hidden lookups
    cp["A4"] = "wizard index"
    cp["B4"] = "=IFERROR(MATCH($B$1,Wizards!$A$2:$A$%d,0),\"\")" % (N_WIZ + 1)
    cp["A5"] = "max spell level"
    cp["B5"] = "=IF($B$4=\"\",\"\",INDEX(Wizards!$D$2:$D$%d,$B$4))" % (N_WIZ + 1)
    cp["A6"] = "order of scribes"
    cp["B6"] = "=IF($B$4=\"\",\"\",INDEX(Wizards!$G$2:$G$%d,$B$4))" % (N_WIZ + 1)
    for r in (4, 5, 6):
        cp.row_dimensions[r].hidden = True

    # Four summary boxes across D1:G2.
    # Order of Scribes (ALPG p.2) changes only the DT rate: ten level 1-4 spells
    # or five level 5-9 spells per 1 DT. GP is unaffected.
    # TRUE, not "x": the Want column is a Google Sheets checkbox, which stores
    # a boolean. COUNTIF/SUMIF match TRUE identically in Excel and Sheets.
    lo = 'COUNTIFS($I${t}:$I${b},TRUE,$A${t}:$A${b},"<=4")'
    hi = 'COUNTIFS($I${t}:$I${b},TRUE,$A${t}:$A${b},">=5")'
    dt = ('=IF($B$6="Yes",CEILING({lo}/10,1)+CEILING({hi}/5,1),{lo}+2*{hi})'
          .format(lo=lo, hi=hi))
    boxes = [
        ("D", "Spells selected", '=COUNTIF($I${t}:$I${b},TRUE)'),
        ("E", "Total GP", '=SUMIF($I${t}:$I${b},TRUE,$G${t}:$G${b})'),
        ("F", "Total DT", dt),
        ("G", "DT remaining", "=$B$2-$F$2"),
    ]
    for col, label, formula in boxes:
        h = cp["%s1" % col]
        h.value = label
        h.font = Font(bold=True, size=9)
        h.alignment = Alignment(horizontal="center", wrap_text=True)
        v = cp["%s2" % col]
        v.value = formula.replace("${t}", str(PLAN_TOP)).replace("${b}", str(last))
        v.font = Font(bold=True, size=13)
        v.alignment = Alignment(horizontal="center")
        v.fill = BOX_FILL
        v.border = BORDER

    header(cp, PLAN_TOP - 1,
           ["Level", "Spell", "School", "Source", "Status", "Available from",
            "GP", "DT", "Want"])
    for i, r in enumerate(spells, start=PLAN_TOP):
        mrow = i - PLAN_TOP + 2                     # matching Matrix row
        cp.cell(row=i, column=1, value=int(r["Level"]))
        cp.cell(row=i, column=2, value=r["Name"])
        cp.cell(row=i, column=3, value=r["School"])
        cp.cell(row=i, column=4, value=r["Source"])
        mine = "INDEX(Matrix!$%s%d:$%s%d,1,$B$4)" % (
            get_column_letter(name_first), mrow,
            get_column_letter(name_last), mrow)
        # OWNED is tested first, so by the time we reach the "# Owners" test the
        # selected wizard is not among the owners -- a count of 0 really does
        # mean nobody else has it.
        cp.cell(row=i, column=5, value=(
            '=IF($B$4="","",'
            'IF({mine}=1,"OWNED",'
            'IF(Spells!$E${srow}<>"","RESTRICTED",'
            'IF($A{row}>$B$5,"TOO HIGH",'
            'IF(Matrix!${cc}{mrow}=0,"NOBODY HAS IT","CAN COPY")))))'
        ).format(mine=mine, srow=mrow, row=i, mrow=mrow,
                 cc=get_column_letter(cnt_col)))
        cp.cell(row=i, column=6, value="=Matrix!$%s%d" % (get_column_letter(own_col), mrow))
        cp.cell(row=i, column=7, value="=$A%d*50" % i)
        cp.cell(row=i, column=8, value="=IF($A%d<=4,1,2)" % i)
        w = cp.cell(row=i, column=9)
        w.fill = PatternFill("solid", fgColor="FFF3B0")
        w.border = BORDER
        w.alignment = Alignment(horizontal="center")
    cp.freeze_panes = "A%d" % PLAN_TOP
    cp.auto_filter.ref = "A%d:I%d" % (PLAN_TOP - 1, last)

    dv_char = DataValidation(type="list",
                             formula1="=Wizards!$A$2:$A$%d" % (N_WIZ + 1),
                             allow_blank=True, showDropDown=False)
    cp.add_data_validation(dv_char)
    dv_char.add(pick)
    # No .xlsx data validation for Want: real checkboxes are applied by
    # build/PolishSpellExchangeSheet.gs once the workbook is in Google Sheets.

    # ---------------------------------------------------------------- Calc
    # One column: for each spell row, its 1-based index if the player ticked
    # Want, otherwise blank. SMALL() over this compacts the ticked spells into
    # a gap-free list on the Copy Log without needing array formulas.
    cal = wb.create_sheet("Calc")
    cal["A1"] = "selected spell index (hidden helper for Copy Log)"
    for i in range(n):
        cal.cell(row=2 + i, column=1,
                 value="=IF('Copy Planner'!$I%d=TRUE,%d,\"\")" % (PLAN_TOP + i, i + 1))
    cal.sheet_state = "hidden"

    # ------------------------------------------------------------ Copy Log
    # Per SPELL, not per lender. Gold and downtime are expended by the copier,
    # so an invoice-shaped table addressed to each lender was simply the wrong
    # model: the wizard doing the copying pays, and what they need is a record
    # of what they bought, what it cost, and who they got it from.
    LOG_LABEL = 8                     # "paste this" banner
    LOG_TEXT = 9                      # the one cell players copy
    LOG_TOP = 12                      # first spell row
    LOG_ROWS = 40                     # a 10 DT budget cannot buy more than this
    ts = wb.create_sheet("Copy Log")
    ts.column_dimensions["A"].width = 5
    ts.column_dimensions["B"].width = 34
    ts.column_dimensions["C"].width = 7
    ts.column_dimensions["D"].width = 9
    ts.column_dimensions["E"].width = 34
    ts.column_dimensions["F"].width = 30
    ts.column_dimensions["G"].width = 3

    ts["A1"] = "COPY LOG"
    ts["A1"].font = TITLE_FONT
    ts["A2"] = ("Everything below is what YOUR selected character spends. "
                "Gold and downtime are expended, not paid to the lender.")
    ts["A2"].font = NOTE_FONT

    for col, label, ref in (("A", "Character", "'Copy Planner'!$B$1"),
                            ("C", "Spells", "'Copy Planner'!$D$2"),
                            ("D", "Total GP", "'Copy Planner'!$E$2"),
                            ("E", "Total DT", "'Copy Planner'!$F$2")):
        h = ts["%s4" % col]
        h.value = label
        h.font = Font(bold=True, size=9)
        v = ts["%s5" % col]
        v.value = "=%s" % ref
        v.font = Font(bold=True, size=13)
        v.fill = BOX_FILL
        v.border = BORDER

    ts["A6"] = ("Downtime is spent as one batch, so Total DT already applies the "
                "Order of Scribes rate where it is due - which is why there is no "
                "per-spell DT column. Per-spell GP is shown, because gold is per "
                "spell level regardless of subclass.")
    ts["A6"].font = NOTE_FONT

    frag_first = 8                                  # hidden per-spell log fragment
    ply_first = frag_first + 1                      # hidden per-wizard player names
    ply_last = ply_first + N_WIZ - 1
    log_rng = "%s%d:%s%d" % (get_column_letter(frag_first), LOG_TOP,
                             get_column_letter(frag_first), LOG_TOP + LOG_ROWS - 1)

    # The single cell players actually copy. It previously sat unlabelled between
    # two notes and the table header, so it read as more commentary and nobody
    # could tell it was the thing to take -- hence the banner directly above.
    banner = ts.cell(row=LOG_LABEL, column=1,
                     value=("PASTE THIS ONE CELL INTO YOUR CHARACTER LOG   "
                            "(click A%d, then copy)" % LOG_TEXT))
    banner.font = HDR_FONT
    banner.fill = HDR_FILL
    for c in range(2, 7):
        ts.cell(row=LOG_LABEL, column=c).fill = HDR_FILL

    txt = ts.cell(row=LOG_TEXT, column=1, value=(
        '=IF(\'Copy Planner\'!$B$4="",'
        '"Pick your character on Copy Planner, then tick the spells you want.",'
        'IF(\'Copy Planner\'!$D$2=0,"Nothing ticked yet.",'
        '"Copied "&\'Copy Planner\'!$D$2&" wizard spell(s) into spellbook: "'
        '&TEXTJOIN("; ",TRUE,{rng})&". Total: "&\'Copy Planner\'!$E$2'
        '&" GP and "&\'Copy Planner\'!$F$2&" DT."))').format(rng=log_rng))
    txt.fill = PatternFill("solid", fgColor="FFF3B0")
    txt.border = BORDER
    txt.alignment = Alignment(wrap_text=True, vertical="top")
    ts.row_dimensions[LOG_TEXT].height = 58

    # Merge both rows across the table's width. The banner then reads as one
    # heading rather than text bleeding over five columns, and the log entry
    # becomes a single obvious click-target -- which is the whole point of it.
    # Merges survive the .xlsx -> Google Sheets conversion, and copying a merged
    # cell still yields its text.
    ts.merge_cells(start_row=LOG_LABEL, start_column=1, end_row=LOG_LABEL, end_column=6)
    ts.merge_cells(start_row=LOG_TEXT, start_column=1, end_row=LOG_TEXT, end_column=6)

    header(ts, LOG_TOP - 1,
           ["#", "Spell", "Level", "GP", "Copied from (character)",
            "Player", "", "log fragment"])
    for c in range(frag_first, ply_last + 1):
        ts.column_dimensions[get_column_letter(c)].hidden = True
    ts.column_dimensions[get_column_letter(frag_first)].width = 60

    own_letter = get_column_letter(own_col)
    for k in range(LOG_ROWS):
        r = LOG_TOP + k
        # k-th ticked spell, compacted out of the Calc column by SMALL().
        ts.cell(row=r, column=1,
                value="=IFERROR(SMALL(Calc!$A$2:$A$%d,%d),\"\")" % (n + 1, k + 1))
        ts.cell(row=r, column=2,
                value="=IF($A%d=\"\",\"\",INDEX(Spells!$B$2:$B$%d,$A%d))" % (r, n + 1, r))
        ts.cell(row=r, column=3,
                value="=IF($A%d=\"\",\"\",INDEX(Spells!$A$2:$A$%d,$A%d))" % (r, n + 1, r))
        ts.cell(row=r, column=4,
                value="=IF($A%d=\"\",\"\",$C%d*50)" % (r, r))
        ts.cell(row=r, column=5,
                value="=IF($A%d=\"\",\"\",INDEX(Matrix!$%s$2:$%s$%d,$A%d))"
                      % (r, own_letter, own_letter, n + 1, r))
        ts.cell(row=r, column=6,
                value="=IF($A%d=\"\",\"\",TEXTJOIN(\", \",TRUE,$%s%d:$%s%d))"
                      % (r, get_column_letter(ply_first), r,
                         get_column_letter(ply_last), r))
        # Guard on $A (the index), never on a numeric column: an unused row
        # returns "" not 0, and testing =0 built stray " (from )" fragments.
        ts.cell(row=r, column=frag_first,
                value=("=IF($A%d=\"\",\"\",$B%d&\" (L\"&$C%d&\", \"&$D%d"
                       "&\" GP) from \"&$E%d)") % (r, r, r, r, r))
        for w in range(1, N_WIZ + 1):
            mcol = get_column_letter(name_first + w - 1)
            ts.cell(row=r, column=ply_first + w - 1,
                    value=("=IF($A%d=\"\",\"\",IF(INDEX(Matrix!$%s$2:$%s$%d,$A%d)=1,"
                           "Wizards!$B$%d,\"\"))")
                          % (r, mcol, mcol, n + 1, r, w + 1))

    ts.cell(row=LOG_TOP + LOG_ROWS + 1, column=1,
            value=("=IF('Copy Planner'!$D$2>%d,\"More than %d spells ticked - "
                   "the list above is truncated.\",\"\")" % (LOG_ROWS, LOG_ROWS))
            ).font = NOTE_FONT
    ts.cell(row=LOG_TOP + LOG_ROWS + 3, column=1,
            value=("AL rule: you may copy from a character's spellbook immediately "
                   "after a session in which you both played - settle it at the event.")
            ).font = NOTE_FONT
    ts.freeze_panes = "A%d" % LOG_TOP

    os.makedirs(os.path.dirname(out), exist_ok=True)
    wb.save(out)
    print("Wrote %s" % out)
    print("  %d spells, %d wizard slots" % (n, N_WIZ))
    print("  tabs: %s" % ", ".join(w.title for w in wb.worksheets))


if __name__ == "__main__":
    build()
